from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def export_game_history(
    output_path: str | Path,
    *,
    game_name: str,
    series: str,
    called_numbers: list[int] | tuple[int, ...],
) -> Path:
    """Exporta una partida en un Excel simple y utilizable por administración."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Partidas"
    headers = ["Fecha y hora", "Juego", "Serie", "Bolas cantadas", "Historial"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    sheet.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        game_name,
        series,
        len(called_numbers),
        ", ".join(str(number) for number in called_numbers),
    ])
    widths = {1: 22, 2: 24, 3: 18, 4: 16, 5: 60}
    for column, width in widths.items():
        sheet.column_dimensions[chr(64 + column)].width = width
    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path
