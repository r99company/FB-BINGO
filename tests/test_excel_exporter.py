from pathlib import Path

from openpyxl import load_workbook

from app.reports.excel_exporter import export_game_history


def test_export_game_history_creates_operational_excel(tmp_path: Path):
    output = tmp_path / "FB-BINGO-reporte.xlsx"
    export_game_history(
        output,
        game_name="Partida rápida",
        series="SERIE-001",
        called_numbers=[47, 12, 90, 5],
        finished_at="2026-09-06 12:34:56",
        status="finalizada",
    )
    assert output.exists()
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Partidas"]
    sheet = workbook["Partidas"]
    assert [sheet.cell(1, column).value for column in range(1, 7)] == ["Fecha y hora", "Juego", "Serie", "Bolas cantadas", "Historial", "Estado"]
    assert sheet.cell(2, 1).value == "2026-09-06 12:34:56"
    assert sheet.cell(2, 2).value == "Partida rápida"
    assert sheet.cell(2, 3).value == "SERIE-001"
    assert sheet.cell(2, 4).value == 4
    assert sheet.cell(2, 5).value == "47, 12, 90, 5"
    assert sheet.cell(2, 6).value == "finalizada"
